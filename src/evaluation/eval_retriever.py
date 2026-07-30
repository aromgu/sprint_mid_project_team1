# -*- coding: utf-8 -*-
"""
eval_retriever.py
=================
골든셋(rag_golden_set.xlsx)으로 하이브리드 리트리버의 검색 성능을 평가한다.

측정 지표 (전부 LLM 불필요 = 무료·결정적)
--------------------------------------------------------------------
  Hit@k        정답 청크가 상위 k개 안에 하나라도 있으면 1
  Recall@k     정답 청크 중 몇 %를 상위 k개 안에서 찾았는가   ← 가장 중요
  MRR@k        첫 정답 청크의 순위 역수 (1등=1.0, 2등=0.5, 3등=0.33)

핵심 설계
--------------------------------------------------------------------
  · 검색은 질문당 딱 1회만 수행하고(k=MAX_K), 결과를 잘라서 k=1/3/5/10/20을
    한꺼번에 계산한다. k마다 검색을 다시 돌리면 20배 느려지기 때문.
  · 검색 결과는 retrieval_results.jsonl 에 저장되므로, 지표를 다시 계산하거나
    k를 바꿔 볼 때 재검색이 필요 없다.
  · 리랭커 유무를 비교할 수 있다 (--stage rerank / hybrid / both).
  · 실행 전에 chunk_id 형식이 골든셋과 맞는지 먼저 검사한다. 
"""

import os
import io
import re
import sys
import json
import time
import math
import argparse
import contextlib
from datetime import datetime

import pandas as pd


# =====================================================================
# CONFIG
# =====================================================================

class CONFIG:
    GOLDEN_XLSX = "rag_golden_set.xlsx"
    GOLDEN_SHEET = "골든셋"
    RESULT_JSONL = "retrieval_results.jsonl"    # 검색 결과 저장 (체크포인트)
    REPORT_XLSX = "retriever_eval.xlsx"

    # 검색은 이 개수만큼 한 번에 가져오고, 지표는 잘라서 계산한다
    MAX_K = 20
    CANDIDATE_K = 30            # search_documents 의 candidate_k (리랭킹 후보 수)
    KS = [1, 3, 5, 10, 20]      # 지표를 계산할 k 목록

    # 리랭커(torch 모델)와 Kiwi는 스레드 안전을 보장하기 어려우므로 기본 1
    # (GPU 메모리에 여유가 있고 안정적이면 2~4로 올려도 됩니다)
    WORKERS = 1

    # ---- 내용 일치(soft match) 보정 ----
    #   나라장터 RFP는 지식재산권·하자보수·청렴서약처럼 똑같은 문구가 여러 사업 문서에
    #   반복 등장한다. 이때 리트리버가 "내용은 맞지만 다른 문서의 청크"를 가져오면
    #   ID가 달라서 오답으로 계산된다(= 실제보다 점수가 낮게 나옴).
    #   그래서 ID가 달라도 텍스트가 거의 같으면 정답으로 인정하는 지표를 함께 계산한다.
    CHUNKS_XLSX = "chroma_chunks.xlsx"   # 청크 원문 (없으면 soft 지표는 생략)
    SOFT_MATCH = True
    SOFT_THRESHOLD = 0.80                # 문자 5-gram 자카드 유사도 기준

    # 실행 이름표. 리트리버/청크버전을 바꿔가며 비교할 때 결과를 구분하는 용도.
    LABEL = ""

    GPU_CACHE_CLEAR_EVERY = 20

    # 골든셋 컬럼명
    COL_ID = "질의ID"
    COL_LEVEL = "레벨명"
    COL_QUESTION = "질문"
    COL_BEHAVIOR = "기대동작"
    COL_SRC_IDS = "근거청크ID"
    COL_DIS_IDS = "방해청크ID"
    COL_NSRC = "근거청크수"
    ID_SEP = " | "


@contextlib.contextmanager
def quiet():
    """
    리트리버가 찍는 로그를 잠시 숨긴다.
    (search_documents 는 질문마다 여러 줄을 출력해서, 900건을 돌리면 화면이 폭발함)
    """
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        yield buf


# =====================================================================
# [0] 골든셋 로드
# =====================================================================

def split_ids(v) -> list:
    """'a | b | c' 셀을 리스트로. 빈 값이면 빈 리스트."""
    if pd.isna(v) or not str(v).strip():
        return []
    return [x.strip() for x in str(v).split(CONFIG.ID_SEP) if x.strip()]


def load_golden(limit: int = 0) -> pd.DataFrame:
    """
    골든셋을 읽고, 검색 평가 대상만 남긴다.

    부정거부(기대동작=refuse) 100건은 '정답 청크'라는 개념 자체가 없으므로
    Recall/MRR 계산에서 제외한다. (그건 생성 단계에서 거절 여부로 평가할 항목)
    """
    df = pd.read_excel(CONFIG.GOLDEN_XLSX, sheet_name=CONFIG.GOLDEN_SHEET)
    df["gold_ids"] = df[CONFIG.COL_SRC_IDS].map(split_ids)
    df["distractor_ids"] = df[CONFIG.COL_DIS_IDS].map(split_ids)
    # 정답 근거의 '원문'. chunk_id가 아니라 내용으로 평가할 때 사용한다.
    # → 청크 사이즈가 다른 DB(512/1024/…)끼리도 같은 골든셋으로 비교 가능해진다.
    if "근거청크원문" in df.columns:
        df["gold_texts"] = df["근거청크원문"].map(
            lambda v: [] if pd.isna(v) else
            [t.strip() for t in str(v).split("\n---\n") if t.strip()])
    else:
        df["gold_texts"] = [[] for _ in range(len(df))]

    total = len(df)
    df = df[(df[CONFIG.COL_BEHAVIOR] == "answer") & (df["gold_ids"].map(len) > 0)].copy()

    print(f"[0] 골든셋: 전체 {total:,}건 → 검색 평가 대상 {len(df):,}건 "
          f"(부정거부 {total - len(df):,}건 제외)")
    for lv, g in df.groupby(CONFIG.COL_LEVEL):
        avg = g["gold_ids"].map(len).mean()
        print(f"      {lv:>10}: {len(g):>4}건 (정답 청크 평균 {avg:.2f}개)")

    if limit and limit < len(df):
        # 레벨 비율을 유지하면서 표본 추출
        df = (df.groupby(CONFIG.COL_LEVEL, group_keys=False)
                .apply(lambda g: g.sample(max(1, round(limit * len(g) / len(df))),
                                          random_state=42)))
        print(f"      → --limit 적용: {len(df):,}건만 평가")
    return df


# =====================================================================
# [1] chunk_id 호환성 사전 점검  ★ 가장 중요한 단계
# =====================================================================

def diagnose_error(e: Exception) -> str:
    """
    검색 중 발생한 예외를 보고, 초보자가 바로 조치할 수 있게 원인을 안내한다.
    """
    name = type(e).__name__
    msg = str(e)

    # 진짜 속도 제한
    if name == "RateLimitError" or "rate_limit" in msg:
        return ("원인: OpenAI 호출 속도 제한(429 rate_limit)입니다.\n"
                "      잠시 후 다시 실행하거나 요청 간격을 늘리세요.")
    if "AuthenticationError" in name or "invalid_api_key" in msg:
        return ("원인: API 키가 잘못되었습니다.\n"
                "      .env 의 OPENAI_API_KEY 값을 확인하세요.")
    if "Collection" in msg and "does not exist" in msg:
        return ("원인: Chroma 컬렉션을 찾을 수 없습니다.\n"
                "      retriever.py 의 CHROMA_COLLECTION / CHROMA_PERSIST_DIR 를 확인하세요.")
    if isinstance(e, FileNotFoundError):
        return ("원인: 파일 경로 문제입니다 (BM25 인덱스 pickle 등).\n"
                "      retriever.py 의 BM25_INDEX_PATH 를 확인하세요.")
    return "원인을 자동으로 판별하지 못했습니다. 위 오류 메시지를 확인하세요."


def preflight_check(df: pd.DataFrame, n_probe: int = 5, stage: str = "rerank") -> bool:
    """
    골든셋의 '근거청크ID'와 리트리버가 돌려주는 chunk_id 형식이 같은지 확인한다.

    이 둘이 다르면 Recall@k, MRR이 전부 0.0으로 나온다.
    리트리버 성능이 나빠서가 아니라 ID가 안 맞아서 생기는 착시인데,
    실제로 가장 흔한 실수라서 본 실행 전에 반드시 확인한다.
    """
    print("\n[1] chunk_id 호환성 점검")

    gold_ids = {i for ids in df["gold_ids"] for i in ids}
    sample_gold = sorted(gold_ids)[:3]
    print(f"    골든셋 정답 청크ID: 고유 {len(gold_ids):,}개")
    for g in sample_gold:
        print(f"      예) {g}")

    # (a) 리트리버를 실제로 몇 번 돌려서 어떤 ID가 나오는지 확인
    try:
        from retriever import search_documents
    except Exception as e:
        print(f"    [실패] retriever.py 를 import 할 수 없습니다: {e}")
        print("\n    " + diagnose_error(e))
        return False

    probe_qs = df[CONFIG.COL_QUESTION].head(n_probe).tolist()
    got_ids = []
    for q in probe_qs:
        try:
            with quiet():
                if stage == "bm25":
                    got = search_bm25_only(str(q), k=5)[0]
                elif stage == "hybrid":
                    got = search_hybrid_only(str(q), k=5)[0]
                else:
                    got = [r.get("id") for r in
                           search_documents(str(q), k=5,
                                            candidate_k=CONFIG.CANDIDATE_K)]
            got_ids += [g for g in got if g]
        except Exception as e:
            print(f"    [실패] 검색 중 오류: {type(e).__name__}: {str(e)[:200]}")
            print("\n    " + diagnose_error(e))
            return False

    if not got_ids:
        print("    [실패] 검색 결과가 비어 있습니다. retriever 설정을 확인하세요.")
        return False

    print(f"\n    리트리버 반환 chunk_id: {len(set(got_ids))}개 (표본 질문 {n_probe}개 기준)")
    for g in list(dict.fromkeys(got_ids))[:3]:
        print(f"      예) {g}")

    # (b) 형식이 겹치는지 판정
    overlap = gold_ids & set(got_ids)
    if overlap:
        print(f"\n    ✅ 정상: 표본 검색에서 정답 청크 {len(overlap)}개가 실제로 검색됨")
        return True

    # 겹치지 않는다면 '형식이 아예 다른지' vs '단지 못 찾은 건지' 구분
    def shape(s):
        """ID의 대략적인 생김새 (콜론 개수, 길이)"""
        return (str(s).count(":"), len(str(s)) // 10)

    gold_shapes = {shape(x) for x in list(gold_ids)[:200]}
    got_shapes = {shape(x) for x in got_ids}

    if gold_shapes & got_shapes:
        print("\n    ⚠ 형식은 같아 보이지만 표본에서는 정답이 안 잡혔습니다.")
        print("      (검색 성능 문제일 수 있으니 그대로 진행해도 됩니다)")
        return True

    print("\n    ❌ 형식 불일치! 이대로 평가하면 모든 점수가 0으로 나옵니다.")
    print("       골든셋 ID 형태:", sorted(gold_shapes))
    print("       리트리버 ID 형태:", sorted(got_shapes))
    print("\n       확인할 것:")
    print("         1) 골든셋을 만든 청크와 지금 Chroma 컬렉션이 같은 버전인가?")
    print("         2) retriever.py 의 metadata['chunk_id'] 값이 골든셋의 '청크ID'와 같은가?")
    print("         3) 컬렉션이 다르면(예: _1024 버전) 그 청크로 골든셋을 다시 만들어야 합니다.")
    return False


# =====================================================================
# [2] 검색 실행 → retrieval_results.jsonl
# =====================================================================

def search_hybrid_only(question: str, k: int) -> list:
    """
    리랭킹 없이 하이브리드(BM25+벡터) 결과만 가져온다.

    ★ 공정한 비교를 위한 핵심 ★
      리트리버에 넘기는 k를 MAX_K가 아니라 CANDIDATE_K로 맞춘다.
      그래야 리랭커가 채점하는 후보 풀과 '똑같은 문서 집합'이 되고,
      두 stage의 차이가 오직 "순서를 매기는 방식"(RRF vs 리랭커)만 남는다.

      만약 여기서 k를 다르게 주면 후보 자체가 달라져서,
      점수 차이가 리랭커 덕분인지 후보가 많아서인지 구분할 수 없게 된다.
    """
    from retriever import get_hybrid_retriever
    retriever = get_hybrid_retriever(k=CONFIG.CANDIDATE_K)
    docs = retriever.invoke(question)[:k]
    return ([d.metadata.get("chunk_id") for d in docs],
            [d.page_content for d in docs])


_bm25 = None


def free_gpu_cache():
    """
    GPU에 쌓인 PyTorch 캐시를 비운다.

    리랭커를 질문마다 반복 호출하면 할당된 메모리 블록이 조각조각 남아
    (fragmentation) '빈 공간은 있는데 큰 덩어리를 못 잡는' 상태가 된다.
    그때 CUDACachingAllocator OOM 경고가 뜬다.
    주기적으로 캐시를 비워주면 경고가 줄고 속도도 안정된다.
    """
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
    except Exception:
        pass          # torch가 없거나 CPU 환경이면 그냥 넘어간다


def search_bm25_only(question: str, k: int) -> list:
    """
    BM25(키워드 검색)만으로 검색한다.

    ★ OpenAI API를 전혀 쓰지 않는다 ★
    벡터 검색은 질문을 임베딩하기 위해 OpenAI API를 호출하므로, API 크레딧이
    없거나 429(quota) 오류가 날 때는 이 모드로 먼저 평가를 진행할 수 있다.
    파이프라인 연결과 chunk_id 매칭이 제대로 되는지 무료로 확인하는 용도.
    """
    global _bm25
    if _bm25 is None:
        from retriever import build_bm25_retriever, _get_cached_documents
        _bm25 = build_bm25_retriever(_get_cached_documents(), k=k)
    _bm25.k = k
    docs = _bm25.invoke(question)[:k]
    return ([d.metadata.get("chunk_id") for d in docs],
            [d.page_content for d in docs])


def run_search(df: pd.DataFrame, stage: str):
    """
    모든 질문을 리트리버에 통과시켜 결과 ID 목록을 저장한다.

    stage:
      "rerank" — search_documents() 전체 파이프라인 (기본, 실제 서비스와 동일)
      "hybrid" — 리랭킹 전 하이브리드 결과만
      "both"   — 둘 다 (리랭커 효과 비교용)
    """
    from retriever import search_documents

    # variant = stage + candidate_k 조합.
    # candidate_k 를 바꿔가며 비교할 수 있도록 설정값을 결과에 함께 기록한다.
    # (이게 없으면 candidate_k=10 으로 다시 돌려도 예전 30 결과를 재사용해버림)
    def variant_of(s):
        # hybrid 도 후보 풀 크기(candidate_k)에 따라 결과가 달라지므로 함께 표기
        base = s if s == "bm25" else f"{s}_c{CONFIG.CANDIDATE_K}"
        return f"{CONFIG.LABEL}|{base}" if CONFIG.LABEL else base

    done = set()
    if os.path.exists(CONFIG.RESULT_JSONL):
        with open(CONFIG.RESULT_JSONL, encoding="utf-8") as f:
            for line in f:
                try:
                    r = json.loads(line)
                    done.add((r["qid"], r.get("variant", r["stage"])))
                except json.JSONDecodeError:
                    pass
        if done:
            print(f"[2] 기존 검색 결과 {len(done):,}건 발견 → 건너뜁니다")

    stages = ["rerank", "hybrid"] if stage == "both" else [stage]
    if stage == "bm25":
        stages = ["bm25"]
    todo = [(i, r, s) for s in stages
            for i, r in df.iterrows()
            if (r[CONFIG.COL_ID], variant_of(s)) not in done]

    if not todo:
        print("[2] 새로 검색할 질문이 없습니다")
        return

    print(f"[2] 검색 실행: {len(todo):,}건 "
          f"(stage={stage}, k={CONFIG.MAX_K}, candidate_k={CONFIG.CANDIDATE_K}"
          + (f", label={CONFIG.LABEL}" if CONFIG.LABEL else "") + ")")
    # EnsembleRetriever는 BM25 k개 + 벡터 k개를 RRF로 합치므로,
    # 리랭커가 실제로 채점하는 후보는 candidate_k가 아니라 '최대 2배'가 된다.
    print(f"    ※ 리랭커가 질문마다 후보 최대 {CONFIG.CANDIDATE_K * 2}개"
          f"(BM25 {CONFIG.CANDIDATE_K} + 벡터 {CONFIG.CANDIDATE_K}, 중복 제거)를 "
          f"채점하므로 시간이 걸립니다")

    t0 = time.time()
    with open(CONFIG.RESULT_JSONL, "a", encoding="utf-8") as f:
        for n, (_, row, s) in enumerate(todo, 1):
            q = str(row[CONFIG.COL_QUESTION])
            t1 = time.time()
            try:
                if s == "rerank":
                    with quiet():
                        res = search_documents(q, k=CONFIG.MAX_K,
                                               candidate_k=CONFIG.CANDIDATE_K)
                    ids = [r.get("id") for r in res]
                    texts = [str(r.get("text") or "") for r in res]
                    scores = [r.get("score") for r in res]
                elif s == "bm25":
                    with quiet():
                        ids, texts = search_bm25_only(q, k=CONFIG.MAX_K)
                    scores = [None] * len(ids)
                else:
                    with quiet():
                        ids, texts = search_hybrid_only(q, k=CONFIG.MAX_K)
                    scores = [None] * len(ids)
                err = None
            except Exception as e:
                ids, texts, scores, err = [], [], [], f"{type(e).__name__}: {str(e)[:200]}"

            f.write(json.dumps({
                "qid": row[CONFIG.COL_ID],
                "stage": s,
                "variant": variant_of(s),     # 비교용 라벨 (예: rerank_c10)
                "candidate_k": CONFIG.CANDIDATE_K,
                "question": q,
                "retrieved_ids": ids,
                # 청크 원문도 저장 → 청크 사이즈가 다른 DB끼리도 '내용'으로 비교 가능
                "retrieved_texts": [t[:3000] for t in texts],
                "scores": scores,
                "latency_sec": round(time.time() - t1, 3),
                "error": err,
            }, ensure_ascii=False) + "\n")
            f.flush()   # 중간에 끊겨도 여기까지는 보존

            # 일정 주기로 GPU 캐시 정리 (OOM 경고 완화)
            if n % CONFIG.GPU_CACHE_CLEAR_EVERY == 0:
                free_gpu_cache()

            if n % 10 == 0 or n == len(todo):
                el = time.time() - t0
                eta = el / n * (len(todo) - n)
                print(f"    {n:,}/{len(todo):,} ({n/len(todo):.0%}) | "
                      f"경과 {el/60:.1f}분 | 예상 잔여 {eta/60:.1f}분 | "
                      f"질문당 {el/n:.2f}초")

    free_gpu_cache()
    print(f"[2] 완료 → {CONFIG.RESULT_JSONL}")


def load_results() -> pd.DataFrame:
    """저장된 검색 결과를 읽는다."""
    if not os.path.exists(CONFIG.RESULT_JSONL):
        raise FileNotFoundError(
            f"'{CONFIG.RESULT_JSONL}' 이 없습니다. 먼저 --search 를 실행하세요.")
    rows = []
    with open(CONFIG.RESULT_JSONL, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    r = pd.DataFrame(rows)
    # 예전 형식(variant 없음)도 읽을 수 있게 보정
    if "variant" not in r.columns:
        r["variant"] = r["stage"]
    r["variant"] = r["variant"].fillna(r["stage"])
    r = r.drop_duplicates(subset=["qid", "variant"], keep="last")

    n_err = r["error"].notna().sum()
    print(f"[*] 검색 결과 로드: {len(r):,}건" + (f" (에러 {n_err}건)" if n_err else ""))
    for v, g in r.groupby("variant"):
        print(f"      {v:>16}: {len(g):,}건")
    return r


# =====================================================================
# [3] 지표 계산
# =====================================================================

_shingle_cache = {}
_chunk_texts = {}


def load_chunk_texts() -> dict:
    """soft match용 chunk_id → 원문 사전. 파일이 없으면 빈 dict."""
    global _chunk_texts
    if _chunk_texts:
        return _chunk_texts
    if not (CONFIG.SOFT_MATCH and os.path.exists(CONFIG.CHUNKS_XLSX)):
        return {}
    c = pd.read_excel(CONFIG.CHUNKS_XLSX)
    _chunk_texts = dict(zip(c["청크ID"].astype(str).str.strip(),
                            c["청크원문"].astype(str)))
    print(f"[*] soft match용 청크 원문 로드: {len(_chunk_texts):,}개 "
          f"(임계값 {CONFIG.SOFT_THRESHOLD})")
    return _chunk_texts


def _shingles(cid: str) -> set:
    """청크 원문을 문자 5-gram 집합으로 변환 (한 번 계산하면 캐싱)."""
    if cid in _shingle_cache:
        return _shingle_cache[cid]
    txt = _chunk_texts.get(cid, "")
    # 공백·기호를 제거해 표 서식 차이에 흔들리지 않게 정규화
    s = re.sub(r"[^0-9A-Za-z가-힣]", "", txt)
    out = {s[i:i + 5] for i in range(0, len(s) - 4)} if len(s) >= 5 else set()
    _shingle_cache[cid] = out
    return out


def soft_equal(a: str, b: str) -> bool:
    """두 청크의 내용이 사실상 같은지 (자카드 유사도 기준)."""
    sa, sb = _shingles(a), _shingles(b)
    if not sa or not sb:
        return False
    inter = len(sa & sb)
    return inter / (len(sa) + len(sb) - inter) >= CONFIG.SOFT_THRESHOLD


def expand_hits(gold: list, retrieved: list) -> set:
    """
    검색 결과 중 '정답으로 인정할 수 있는' 청크 ID 집합을 만든다.
    - ID가 정확히 같으면 당연히 포함
    - ID는 다르지만 내용이 거의 같으면 함께 포함 (soft match)
    """
    ok = set(gold)
    if not _chunk_texts:
        return ok
    for cid in retrieved:
        if cid in ok:
            continue
        for g in gold:
            if soft_equal(cid, g):
                ok.add(cid)
                break
    return ok


def _ngrams(text: str, n: int = 5) -> set:
    """정규화한 텍스트를 문자 n-gram 집합으로. (표 서식/공백 차이를 무시하기 위함)"""
    t = re.sub(r"[^0-9A-Za-z가-힣]", "", str(text))
    return {t[i:i + n] for i in range(0, len(t) - n + 1)} if len(t) >= n else set()


def coverage_at_k(gold_texts: list, retrieved_texts: list, k: int) -> dict:
    """
    '내용 커버리지' 지표 — chunk_id 대신 텍스트로 정답 여부를 판정한다.

    왜 필요한가
      청크 사이즈가 다른 DB(512 / 1024 / 256 …)는 chunk_id 체계가 완전히 다르다.
      그래서 ID 기반 Recall로는 서로 비교할 수 없다.
      대신 "정답 근거 문장이 검색 결과 안에 얼마나 들어 있는가"를 보면
      청크를 어떻게 쪼갰든 공정하게 비교할 수 있다.

    계산
      정답 청크의 문자 5-gram 중 몇 %가 검색된 청크들의 5-gram 합집합에 있는지.
      1024 청크가 512 정답을 통째로 품고 있으면 커버리지 1.0,
      256 청크가 절반만 담고 있으면 0.5 식으로 자연스럽게 부분점수가 매겨진다.
    """
    if not gold_texts:
        return {}
    pool = set()
    for t in list(retrieved_texts)[:k]:
        pool |= _ngrams(t)

    covs = []
    for g in gold_texts:
        gg = _ngrams(g)
        covs.append(len(gg & pool) / len(gg) if gg else 0.0)

    mean_cov = sum(covs) / len(covs)
    return {
        f"Coverage@{k}": mean_cov,                                   # 평균 커버리지
        f"CovHit@{k}": 1.0 if max(covs) >= 0.6 else 0.0,             # 하나라도 60% 이상
        f"CovAll@{k}": 1.0 if min(covs) >= 0.6 else 0.0,             # 전부 60% 이상
    }


def metrics_at_k(gold: list, retrieved: list, k: int) -> dict:
    """
    한 질문에 대해 상위 k개 기준 지표를 계산한다.

    gold      : 정답 청크 ID 리스트 (골든셋의 근거청크ID)
    retrieved : 리트리버가 순위대로 돌려준 청크 ID 리스트
    """
    gold_set = set(gold)
    top = list(retrieved)[:k]
    hits = [cid for cid in top if cid in gold_set]

    # MRR: 첫 정답이 나온 순위의 역수
    rr = 0.0
    for rank, cid in enumerate(top, start=1):
        if cid in gold_set:
            rr = 1.0 / rank
            break

    # nDCG: 정답이 앞쪽에 있을수록 높음
    dcg = sum(1.0 / math.log2(rank + 1)
              for rank, cid in enumerate(top, start=1) if cid in gold_set)
    ideal_n = min(len(gold_set), k)
    idcg = sum(1.0 / math.log2(rank + 1) for rank in range(1, ideal_n + 1))

    return {
        f"Hit@{k}": 1.0 if hits else 0.0,
        f"Recall@{k}": len(hits) / len(gold_set) if gold_set else 0.0,
        f"Precision@{k}": len(hits) / len(top) if top else 0.0,
        f"MRR@{k}": rr,
        f"nDCG@{k}": (dcg / idcg) if idcg > 0 else 0.0,
    }


def score_all(df: pd.DataFrame, res: pd.DataFrame, ks: list) -> pd.DataFrame:
    """모든 질문 × 모든 k에 대해 지표를 계산해 하나의 표로 만든다."""
    load_chunk_texts()          # soft match 준비 (파일 없으면 자동 생략)
    m = df.merge(res, left_on=CONFIG.COL_ID, right_on="qid", how="inner")

    rows = []
    for _, r in m.iterrows():
        rec = {
            "질의ID": r[CONFIG.COL_ID],
            "stage": r["variant"],          # 예: rerank_c10 / rerank_c30 / hybrid / bm25
            "candidate_k": r.get("candidate_k"),
            "레벨명": r[CONFIG.COL_LEVEL],
            "정답청크수": len(r["gold_ids"]),
            "검색청크수": len(r["retrieved_ids"]),
            "지연시간초": r.get("latency_sec"),
            "질문": r[CONFIG.COL_QUESTION],
            "정답청크ID": CONFIG.ID_SEP.join(r["gold_ids"]),
            "검색청크ID(상위10)": CONFIG.ID_SEP.join(list(r["retrieved_ids"])[:10]),
        }
        for k in ks:
            rec.update(metrics_at_k(r["gold_ids"], r["retrieved_ids"], k))

        # 내용 커버리지 (청크 사이즈가 다른 DB 간 비교용)
        gtexts = r.get("gold_texts") or []
        rtexts = r.get("retrieved_texts") or []
        if gtexts and len(rtexts):
            for k in ks:
                rec.update(coverage_at_k(gtexts, rtexts, k))
                # 상위 k개를 LLM에 넣을 때의 분량. 청크가 클수록 커버리지가 유리해지므로
                # "얼마나 많은 글자를 읽혀서 그 커버리지를 얻었는지"를 함께 봐야 공정하다.
                rec[f"검색문자수@{k}"] = sum(len(str(t)) for t in list(rtexts)[:k])

        # soft match: ID는 달라도 내용이 같은 청크까지 정답으로 인정한 지표
        if _chunk_texts:
            accept = expand_hits(r["gold_ids"], list(r["retrieved_ids"]))
            n_gold = len(r["gold_ids"])
            for k in ks:
                top = list(r["retrieved_ids"])[:k]
                hits = [c for c in top if c in accept]
                rec[f"Hit_soft@{k}"] = 1.0 if hits else 0.0
                # 정답 개수를 넘지 않도록 상한 처리 (중복 청크가 여러 개 잡힐 수 있음)
                rec[f"Recall_soft@{k}"] = (min(len(hits), n_gold) / n_gold) if n_gold else 0.0
            rec["내용일치추가"] = len(accept) - n_gold   # soft로 늘어난 개수
        # 첫 정답이 몇 등에서 나왔는지 (실패 분석용)
        first = next((i for i, c in enumerate(r["retrieved_ids"], 1)
                      if c in set(r["gold_ids"])), None)
        rec["첫정답순위"] = first if first else -1
        rows.append(rec)

    return pd.DataFrame(rows)


def print_summary(scored: pd.DataFrame, ks: list):
    """콘솔에 요약표를 출력한다."""
    metric_names = ["Hit", "Recall", "Precision", "MRR", "nDCG"]

    for stage, g in scored.groupby("stage"):
        core = stage.split("|")[-1]      # 라벨 제거 후 stage 판별
        if core.startswith("rerank"):
            ck = core.split("_c")[-1] if "_c" in core else "?"
            label = f"리랭킹 후 (실제 파이프라인, candidate_k={ck})"
        elif core == "bm25":
            label = "BM25 단독 (키워드 검색만)"
        else:
            ck = core.split("_c")[-1] if "_c" in core else "?"
            label = f"리랭킹 전 (하이브리드 RRF만, candidate_k={ck})"
        print(f"\n{'='*66}")
        print(f" [{stage}] {label}  —  대상 {len(g):,}건")
        print("=" * 66)

        # k별 전체 성능
        print(f"\n {'k':>4} " + "".join(f"{n:>11}" for n in metric_names))
        print(" " + "-" * 60)
        for k in ks:
            vals = [g[f"{n}@{k}"].mean() for n in metric_names]
            print(f" {k:>4} " + "".join(f"{v:>11.3f}" for v in vals))

        # 내용 커버리지 (청크 버전 간 비교용)
        if f"Coverage@{ks[-1]}" in g.columns and g[f"Coverage@{ks[-1]}"].notna().any():
            print(f"\n {'k':>4} {'Coverage':>11}{'CovHit':>9}{'CovAll':>9}   "
                  f"(청크 사이즈 무관 비교용)")
            print(" " + "-" * 48)
            for k in ks:
                chars = g[f"검색문자수@{k}"].mean() if f"검색문자수@{k}" in g.columns else 0
                print(f" {k:>4} {g[f'Coverage@{k}'].mean():>11.3f}"
                      f"{g[f'CovHit@{k}'].mean():>9.3f}{g[f'CovAll@{k}'].mean():>9.3f}"
                      f"   (평균 {chars:,.0f}자 투입)")

        # 내용 일치(soft) 보정 결과
        if f"Recall_soft@{ks[-1]}" in g.columns:
            print(f"\n {'k':>4} {'Recall(엄격)':>13}{'Recall(내용일치)':>16}{'차이':>9}")
            print(" " + "-" * 44)
            for k in ks:
                a, b = g[f"Recall@{k}"].mean(), g[f"Recall_soft@{k}"].mean()
                print(f" {k:>4} {a:>13.3f}{b:>16.3f}{b-a:>+9.3f}")
            n_soft = (g["내용일치추가"] > 0).sum()
            print(f"   ※ 내용이 같은 다른 문서 청크를 가져온 질문: {n_soft}건 "
                  f"({n_soft/len(g):.1%})")

        # 레벨별 (대표 k만)
        rep_k = 5 if 5 in ks else ks[len(ks)//2]
        cols = [f"{n}@{rep_k}" for n in ["Hit", "Recall", "MRR", "nDCG"]]
        print(f"\n ── 레벨별 (k={rep_k}) ──")
        t = g.groupby("레벨명")[cols].mean().round(3)
        t.loc["── 전체"] = g[cols].mean().round(3)
        print(t.to_string())

        # 정답 청크 개수별 — 정답이 여러 개인 질문일수록 Recall이 떨어지는지 확인
        print(f"\n ── 정답 청크 개수별 Recall@{rep_k} ──")
        t2 = (g.groupby("정답청크수")
                .agg(건수=("질의ID", "count"),
                     **{f"Recall@{rep_k}": (f"Recall@{rep_k}", "mean")})
                .round(3))
        print(t2.to_string())

        # 완전 실패 건수
        fail = (g[f"Hit@{max(ks)}"] == 0).sum()
        print(f"\n ── 상위 {max(ks)}개 안에 정답이 하나도 없는 질문: "
              f"{fail:,}건 ({fail/len(g):.1%})")

        if g["지연시간초"].notna().any():
            print(f" ── 질문당 평균 검색 시간: {g['지연시간초'].mean():.2f}초 "
                  f"(중앙값 {g['지연시간초'].median():.2f}초)")

    # 리랭커 효과 비교
    if scored["stage"].nunique() > 1:
        print(f"\n{'='*66}")
        print(" 리랭커 효과 (리랭킹 후 − 리랭킹 전)")
        print("=" * 66)
        piv = scored.groupby("stage")[[f"{n}@{k}" for k in ks
                                       for n in metric_names]].mean()
        rr = [i for i in piv.index if "rerank" in str(i)]
        hy = [i for i in piv.index if str(i).split("|")[-1].startswith("hybrid")]
        if rr and hy:
            for r_name in rr:
                pair = next((h for h in hy
                             if str(h).split("|")[0] == str(r_name).split("|")[0]), hy[0])
                print(f"\n  [{r_name}] vs [{pair}]")
                diff = (piv.loc[r_name] - piv.loc[pair]).round(3)
                for k in ks:
                    line = " | ".join(f"{n}@{k} {diff[f'{n}@{k}']:+.3f}"
                                      for n in ["Recall", "MRR", "nDCG"])
                    print(f"    k={k:<3} {line}")
        # candidate_k 비교 (rerank_c10 vs rerank_c30 처럼 두 개 이상일 때)
        if len(rr) > 1:
            base = sorted(rr)[0]
            print(f"\n  candidate_k 비교 (기준: {base})")
            for other in sorted(rr)[1:]:
                diff = (piv.loc[other] - piv.loc[base]).round(3)
                for k in ks:
                    line = " | ".join(f"{n}@{k} {diff[f'{n}@{k}']:+.3f}"
                                      for n in ["Recall", "MRR", "nDCG"])
                    print(f"    [{other}] k={k:<3} {line}")


# =====================================================================
# [4] 엑셀 리포트
# =====================================================================

def export_report(scored: pd.DataFrame, ks: list, path: str):
    """요약 / 상세 / 실패사례 시트를 담은 엑셀을 만든다."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    metric_names = ["Hit", "Recall", "MRR"]
    # 청크 사이즈가 다른 버전끼리 비교할 때 쓰는 지표 (ID가 아닌 '내용' 기반)
    cov_names = ["Coverage", "CovHit", "CovAll"]

    # --- 요약 시트 ---
    sum_rows = []
    for stage, g in scored.groupby("stage"):
        for k in ks:
            row = {"stage": stage, "k": k, "건수": len(g)}
            for n in metric_names:
                row[n] = round(g[f"{n}@{k}"].mean(), 4)
            for n in cov_names:
                if f"{n}@{k}" in g.columns:
                    row[n] = round(g[f"{n}@{k}"].mean(), 4)
            if f"검색문자수@{k}" in g.columns:
                row["투입문자수"] = int(g[f"검색문자수@{k}"].mean())
            sum_rows.append(row)
    df_sum = pd.DataFrame(sum_rows)

    # --- 레벨별 시트 ---
    lvl_rows = []
    for (stage, lv), g in scored.groupby(["stage", "레벨명"]):
        for k in ks:
            row = {"stage": stage, "레벨명": lv, "k": k, "건수": len(g)}
            for n in metric_names:
                row[n] = round(g[f"{n}@{k}"].mean(), 4)
            for n in cov_names:
                if f"{n}@{k}" in g.columns:
                    row[n] = round(g[f"{n}@{k}"].mean(), 4)
            lvl_rows.append(row)
    df_lvl = pd.DataFrame(lvl_rows)

    # --- 실패 사례 시트 (개선 포인트를 찾는 데 가장 유용) ---
    max_k = max(ks)
    fails = scored[scored[f"Hit@{max_k}"] == 0].copy()
    keep = ["stage", "질의ID", "레벨명", "정답청크수", "질문",
            "정답청크ID", "검색청크ID(상위10)", "지연시간초"]
    df_fail = fails[[c for c in keep if c in fails.columns]]

    # --- 상세 시트 ---
    detail_cols = (["stage", "질의ID", "레벨명", "정답청크수", "첫정답순위", "지연시간초"]
                   + [f"{n}@{k}" for k in ks for n in metric_names]
                   + ["질문", "정답청크ID", "검색청크ID(상위10)"])
    df_det = scored[[c for c in detail_cols if c in scored.columns]]

    df_meta = pd.DataFrame([
        {"항목": "평가일시", "값": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"항목": "골든셋", "값": CONFIG.GOLDEN_XLSX},
        {"항목": "평가 질문 수", "값": scored["질의ID"].nunique()},
        {"항목": "MAX_K(검색 개수)", "값": CONFIG.MAX_K},
        {"항목": "CANDIDATE_K(리랭킹 후보)", "값": CONFIG.CANDIDATE_K},
        {"항목": "계산한 k", "값": ", ".join(map(str, ks))},
    ])

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df_meta.to_excel(w, sheet_name="정보", index=False)
        df_sum.to_excel(w, sheet_name="요약", index=False)
        df_lvl.to_excel(w, sheet_name="레벨별", index=False)
        df_fail.to_excel(w, sheet_name="실패사례", index=False)
        df_det.to_excel(w, sheet_name="상세", index=False)

        wb = w.book
        fill = PatternFill("solid", fgColor="2F5597")
        font = Font(color="FFFFFF", bold=True, size=10)
        for name in ("정보", "요약", "레벨별", "실패사례", "상세"):
            ws = wb[name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for c in ws[1]:
                c.fill, c.font = fill, font
                c.alignment = Alignment(horizontal="center")
            for i in range(1, ws.max_column + 1):
                head = ws.cell(row=1, column=i).value
                width = 55 if head in ("질문",) else (
                    40 if head and "청크ID" in str(head) else 14)
                ws.column_dimensions[get_column_letter(i)].width = width

    print(f"\n[4] 리포트 저장: {os.path.abspath(path)}")
    print(f"    실패 사례 {len(df_fail):,}건이 '실패사례' 시트에 정리돼 있습니다.")


# =====================================================================
# MAIN
# =====================================================================

def main():
    p = argparse.ArgumentParser(description="골든셋으로 리트리버 Recall@k / MRR 평가")
    p.add_argument("--check", action="store_true", help="chunk_id 호환성만 점검하고 종료")
    p.add_argument("--search", action="store_true", help="검색만 실행")
    p.add_argument("--score", action="store_true", help="저장된 결과로 지표만 계산")
    p.add_argument("--stage", default="rerank",
                   choices=["rerank", "hybrid", "both", "bm25"],
                   help="rerank=전체 파이프라인 / hybrid=리랭킹 전 / both=비교 / "
                        "bm25=키워드 검색만 (OpenAI API 불필요, 무료)")
    p.add_argument("--limit", type=int, default=0, help="이 개수만 평가 (빠른 확인용)")
    p.add_argument("--max-k", type=int, default=CONFIG.MAX_K, help="검색해올 청크 개수")
    p.add_argument("--candidate-k", type=int, default=CONFIG.CANDIDATE_K,
                   help="리랭킹 후보 개수")
    p.add_argument("--ks", default=",".join(map(str, CONFIG.KS)),
                   help="지표를 계산할 k 목록 (예: 1,3,5,10)")
    p.add_argument("--golden", default=CONFIG.GOLDEN_XLSX)
    p.add_argument("--golden-map", default=None,
                   help="라벨별로 다른 골든셋을 쓸 때. 청크 버전마다 chunk_id 체계가 "
                        "다르므로 각자의 정답지로 채점해야 한다. "
                        '예: "V512=golden_150_512.xlsx,V1024=golden_150_1024.xlsx". '
                        '라벨 끝에 *를 붙이면 접두어 매칭 (예: "V1024*=golden_150_1024.xlsx" '
                        '→ V1024, V1024_w55, V1024_w73 을 모두 포함)')
    p.add_argument("--output", default=CONFIG.REPORT_XLSX)
    p.add_argument("--label", default="",
                   help="이번 실행에 이름표를 붙인다 (예: A_512, B_1024). "
                        "라벨이 다르면 결과가 섞이지 않고, 나중에 --score 한 번으로 "
                        "여러 설정을 한 표에서 비교할 수 있다")
    p.add_argument("--results", default=CONFIG.RESULT_JSONL,
                   help="검색 결과 저장 파일 경로 (설정별로 완전히 분리하고 싶을 때)")
    p.add_argument("--fresh", action="store_true", help="기존 검색 결과를 지우고 다시 검색")
    args = p.parse_args()

    CONFIG.GOLDEN_XLSX = args.golden
    CONFIG.REPORT_XLSX = args.output
    CONFIG.MAX_K = args.max_k
    CONFIG.CANDIDATE_K = args.candidate_k
    CONFIG.RESULT_JSONL = args.results
    CONFIG.LABEL = args.label
    ks = sorted({int(x) for x in args.ks.split(",") if x.strip()})
    ks = [k for k in ks if k <= CONFIG.MAX_K] or [CONFIG.MAX_K]

    if args.fresh and os.path.exists(CONFIG.RESULT_JSONL):
        os.remove(CONFIG.RESULT_JSONL)
        print("[*] 기존 검색 결과를 삭제했습니다")

    print("=" * 66)
    print(" 리트리버 검색 성능 평가 (Recall@k / MRR)")
    if CONFIG.LABEL:
        print(f" 라벨: {CONFIG.LABEL}")
    print(f" 결과 파일: {CONFIG.RESULT_JSONL}")
    print("=" * 66)

    # --golden-map 을 쓰면 라벨마다 골든셋이 다르므로, 초기 로드는 첫 번째 파일로 한다
    if args.golden_map:
        first = args.golden_map.split(",")[0]
        if "=" in first:
            CONFIG.GOLDEN_XLSX = first.split("=", 1)[1].strip()

    df = load_golden(limit=args.limit)

    do_all = not (args.check or args.search or args.score)

    if args.check or args.search or do_all:
        probe_stage = "rerank" if args.stage == "both" else args.stage
        ok = preflight_check(df, stage=probe_stage)
        if not ok and not args.check:
            print("\n중단합니다. ID 문제를 먼저 해결하세요. "
                  "(그래도 진행하려면 --search 로 직접 실행)")
            return
        if args.check:
            return

    if args.search or do_all:
        run_search(df, args.stage)

    if args.score or do_all:
        res = load_results()

        if args.golden_map:
            # 라벨(=청크 버전)마다 자기 골든셋으로 따로 채점한 뒤 하나로 합친다.
            # 이렇게 해야 버전별 Recall/MRR이 정상적으로 나온다.
            parts = []
            matched = set()
            for pair in args.golden_map.split(","):
                if "=" not in pair:
                    continue
                label, path = pair.split("=", 1)
                label, path = label.strip(), path.strip()

                # 라벨 끝에 * 가 있으면 접두어 매칭, 없으면 정확히 그 라벨만
                if label.endswith("*"):
                    prefix = label[:-1]
                    mask = res["variant"].astype(str).str.startswith(prefix)
                else:
                    mask = res["variant"].astype(str).str.startswith(label + "|")

                sub = res[mask]
                if len(sub) == 0:
                    print(f"    ⚠ 라벨 '{label}' 에 해당하는 검색 결과가 없습니다")
                    continue
                matched |= set(sub["variant"].unique())

                CONFIG.GOLDEN_XLSX = path
                gdf = load_golden(limit=args.limit)
                found = ", ".join(sorted(sub["variant"].unique()))
                print(f"    [{label}] {os.path.basename(path)} 로 채점 "
                      f"({len(sub)}건) → {found}")
                parts.append(score_all(gdf, sub, ks))

            # 어느 매핑에도 걸리지 않은 변형이 있으면 알려준다 (조용히 빠지는 사고 방지)
            missed = sorted(set(res["variant"].unique()) - matched)
            if missed:
                print(f"\n    ⚠ 매핑에 없어 제외된 변형 {len(missed)}개:")
                for m in missed:
                    print(f"        {m}")
                print("      → --golden-map 에 해당 라벨을 추가하거나 라벨 끝에 * 를 쓰세요")

            if not parts:
                print("계산할 결과가 없습니다.")
                return
            scored = pd.concat(parts, ignore_index=True)
        else:
            scored = score_all(df, res, ks)

        if len(scored) == 0:
            print("계산할 결과가 없습니다.")
            return
        print_summary(scored, ks)
        export_report(scored, ks, CONFIG.REPORT_XLSX)

    print("=" * 66)


if __name__ == "__main__":
    main()